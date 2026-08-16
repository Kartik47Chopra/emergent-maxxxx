from dotenv import load_dotenv
load_dotenv()

import os
import uuid
import logging
from contextlib import asynccontextmanager
from datetime import datetime, timezone, timedelta
import asyncio
import io
import json
from typing import Optional, List, Dict

import bcrypt
import jwt
import requests
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field
from emergentintegrations.llm.chat import LlmChat, UserMessage, TextDelta, StreamDone

STORAGE_BASE = (os.environ.get("INTEGRATION_PROXY_URL") or "").strip() or "https://integrations.emergentagent.com"
STORAGE_URL = STORAGE_BASE.rstrip("/") + "/objstore/api/v1/storage"
APP_NAME = "maxx-doors"
storage_key = None


def init_storage(force: bool = False):
    global storage_key
    if storage_key and not force:
        return storage_key
    resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": os.environ.get("EMERGENT_LLM_KEY")}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key


def put_object(path: str, data: bytes, content_type: str) -> dict:
    resp = requests.put(f"{STORAGE_URL}/objects/{path}",
                        headers={"X-Storage-Key": init_storage(), "Content-Type": content_type},
                        data=data, timeout=120)
    resp.raise_for_status()
    return resp.json()


def get_object(path: str):
    resp = requests.get(f"{STORAGE_URL}/objects/{path}",
                        headers={"X-Storage-Key": init_storage()}, timeout=60)
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

JWT_ALGORITHM = "HS256"
STATIONS = ["core", "skin", "assembly", "press", "routing"]
PREV = {"core": None, "skin": None, "assembly": ["core", "skin"], "press": ["assembly"], "routing": ["press"]}

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))


def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "type": "access",
               "exp": datetime.now(timezone.utc) + timedelta(hours=12)}
    return jwt.encode(payload, os.environ["JWT_SECRET"], algorithm=JWT_ALGORITHM)


def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "type": "refresh",
               "exp": datetime.now(timezone.utc) + timedelta(days=7)}
    return jwt.encode(payload, os.environ["JWT_SECRET"], algorithm=JWT_ALGORITHM)


def set_auth_cookies(response: Response, user_id: str, email: str):
    response.set_cookie("access_token", create_access_token(user_id, email),
                        httponly=True, secure=True, samesite="none", max_age=43200, path="/")
    response.set_cookie("refresh_token", create_refresh_token(user_id),
                        httponly=True, secure=True, samesite="none", max_age=604800, path="/")


def public_user(u: dict) -> dict:
    return {"id": u["id"], "email": u["email"], "name": u["name"],
            "role": u["role"], "station": u.get("station")}


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, os.environ["JWT_SECRET"], algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(401, "Invalid token type")
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
    if not user:
        raise HTTPException(401, "User not found")
    return user


def require_office(user=Depends(get_current_user)):
    if user["role"] != "office":
        raise HTTPException(403, "Office staff only")
    return user


async def seed_users():
    office_email = os.environ.get("ADMIN_EMAIL", "office@maxxdoors.com").lower()
    office_pw = os.environ.get("ADMIN_PASSWORD", "MaxxOffice!2026")
    existing = await db.users.find_one({"email": office_email})
    if not existing:
        await db.users.insert_one({"id": str(uuid.uuid4()), "email": office_email,
                                   "password_hash": hash_password(office_pw), "name": "Office Control",
                                   "role": "office", "station": None, "created_at": now_iso()})
    elif not verify_password(office_pw, existing["password_hash"]):
        await db.users.update_one({"email": office_email},
                                  {"$set": {"password_hash": hash_password(office_pw)}})
    for st in STATIONS:
        email = f"{st}@maxxdoors.com"
        if not await db.users.find_one({"email": email}):
            await db.users.insert_one({"id": str(uuid.uuid4()), "email": email,
                                       "password_hash": hash_password("MaxxStation!2026"),
                                       "name": f"{st.title()} Operator", "role": "operator",
                                       "station": st, "created_at": now_iso()})


def fresh_stages(core_qty: int = 1):
    def s():
        return {"status": "awaiting", "by": None, "at": None}
    stages = {k: s() for k in ["core", "skin", "assembly", "press", "routing", "despatch"]}
    stages["assembly"]["photo"] = None
    stages["routing"]["qc"] = None
    stages["routing"]["notes"] = ""
    if not core_qty:
        stages["core"].update({"status": "completed", "by": "AUTO (no core required)", "at": now_iso()})
    return stages


def make_door(job_id, job_name, floor, door_id, location, door_type="DSC-03d", **kw):
    core_qty = kw.get("core_qty", 1)
    return {
        "id": str(uuid.uuid4()), "job_id": job_id, "job_name": job_name,
        "floor": floor, "location": location, "door_type": door_type,
        "qty": kw.get("qty", 1), "internal_door": kw.get("internal_door", "Yes"),
        "door_id": door_id, "leaf_height": kw.get("leaf_height", "2400"),
        "leaf_width_1": kw.get("leaf_width_1", "826"), "leaf_width_2": kw.get("leaf_width_2", ""),
        "panel_thickness": kw.get("panel_thickness", "54"),
        "actual_thickness": kw.get("actual_thickness", "56"),
        "leaf_type": kw.get("leaf_type", "Single"), "panel_finish": kw.get("panel_finish", "Sapele Veneer"),
        "fire_rating": kw.get("fire_rating", "FD60"),
        "core_qty": core_qty, "core_type": kw.get("core_type", "FR Particleboard 44mm"),
        "core_cutting": kw.get("core_cutting", "2400 x 826"),
        "skin_qty": kw.get("skin_qty", 2), "skin_type": kw.get("skin_type", "Sapele Veneer 6mm"),
        "skin_cutting": kw.get("skin_cutting", "2440 x 840"),
        "extras": kw.get("extras", {}),
        "stages": fresh_stages(core_qty), "created_at": now_iso(),
    }


def set_stage(door, station, by="Seeder"):
    door["stages"][station].update({"status": "completed", "by": by, "at": now_iso()})


async def seed_demo():
    if await db.jobs.count_documents({}) > 0:
        return
    job = {"id": str(uuid.uuid4()), "name": "Riverside Gate - Tower A", "client": "Meridian Construction",
           "released": True, "created_at": now_iso()}
    await db.jobs.insert_one(job)
    locs = ["CORRIDOR ENTRY", "BOH AREA", "FIRE CONTROL ROOM", "STAIR CORE", "RISER CUPBOARD"]
    doors = []
    for i in range(2, 10):
        doors.append(make_door(job["id"], job["name"], "Ground Floor", f"RG01.D{i}", locs[i % len(locs)]))
    for i in range(1, 5):
        doors.append(make_door(job["id"], job["name"], "Level 1", f"RG02.D{i}", locs[i % len(locs)],
                               core_cutting="2100 x 926", skin_cutting="2140 x 940"))
    for d in doors:
        n = int(d["door_id"].split(".D")[1])
        if d["floor"] == "Ground Floor":
            if n <= 4:
                for st in ["core", "skin", "assembly", "press", "routing"]:
                    set_stage(d, st)
                d["stages"]["routing"]["qc"] = "pass"
                if n == 2:
                    set_stage(d, "despatch")
            elif n == 5:
                for st in ["core", "skin", "assembly", "press"]:
                    set_stage(d, st)
            elif n == 6:
                for st in ["core", "skin"]:
                    set_stage(d, st)
            elif n == 7:
                set_stage(d, "core")
        await db.doors.insert_one(d)
    logger.info("Seeded demo job with %d doors", len(doors))


@asynccontextmanager
async def lifespan(app: FastAPI):
    await db.users.create_index("email", unique=True)
    await db.doors.create_index("door_id", unique=True)
    await db.doors.create_index("job_id")
    await db.doors.create_index("floor")
    await seed_users()
    await seed_demo()
    yield
    client.close()


app = FastAPI(lifespan=lifespan)
api_router = APIRouter(prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[os.environ.get("FRONTEND_URL", "http://localhost:3000")],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------- Auth ----------

class LoginIn(BaseModel):
    email: str
    password: str


@api_router.post("/auth/login")
async def login(body: LoginIn, request: Request, response: Response):
    email = body.email.lower().strip()
    ident = f"{request.client.host}:{email}"
    attempts = await db.login_attempts.find_one({"identifier": ident})
    if attempts and attempts.get("count", 0) >= 5:
        locked_at = datetime.fromisoformat(attempts["locked_at"])
        if datetime.now(timezone.utc) - locked_at < timedelta(minutes=15):
            raise HTTPException(429, "Too many failed attempts. Try again in 15 minutes.")
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": ident},
            {"$inc": {"count": 1}, "$set": {"locked_at": now_iso()}}, upsert=True)
        raise HTTPException(401, "Invalid email or password")
    await db.login_attempts.delete_one({"identifier": ident})
    set_auth_cookies(response, user["id"], user["email"])
    return public_user(user)


@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"ok": True}


@api_router.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return public_user(user)


@api_router.post("/auth/refresh")
async def refresh(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(401, "No refresh token")
    try:
        payload = jwt.decode(token, os.environ["JWT_SECRET"], algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(401, "Invalid token type")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid refresh token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
    if not user:
        raise HTTPException(401, "User not found")
    set_auth_cookies(response, user["id"], user["email"])
    return public_user(user)


# ---------- Jobs & Doors ----------

class DoorIn(BaseModel):
    floor: str
    location: str
    door_type: str = "DSC-03d"
    qty: int = 1
    internal_door: str = "Yes"
    door_id: str
    leaf_height: str = ""
    leaf_width_1: str = ""
    leaf_width_2: str = ""
    panel_thickness: str = ""
    actual_thickness: str = ""
    leaf_type: str = "Single"
    panel_finish: str = ""
    fire_rating: str = ""
    core_qty: int = 1
    core_type: str = ""
    core_cutting: str = ""
    skin_qty: int = 2
    skin_type: str = ""
    skin_cutting: str = ""


class JobIn(BaseModel):
    name: str
    client: str = ""
    doors: List[DoorIn] = []


@api_router.get("/jobs")
async def list_jobs(user=Depends(get_current_user)):
    jobs = await db.jobs.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)
    for j in jobs:
        j["door_count"] = await db.doors.count_documents({"job_id": j["id"]})
    return jobs


@api_router.post("/jobs")
async def create_job(body: JobIn, user=Depends(require_office)):
    job = {"id": str(uuid.uuid4()), "name": body.name, "client": body.client,
           "released": False, "created_at": now_iso()}
    ids = [d.door_id.strip() for d in body.doors]
    if len(ids) != len(set(ids)):
        raise HTTPException(400, "Duplicate Door IDs in this job")
    existing = await db.doors.find({"door_id": {"$in": ids}}, {"_id": 0, "door_id": 1}).to_list(1000)
    if existing:
        raise HTTPException(400, "Door IDs already exist: " + ", ".join(e["door_id"] for e in existing))
    await db.jobs.insert_one(job)
    for d in body.doors:
        data = d.model_dump()
        door = make_door(job["id"], job["name"], data.pop("floor"), data.pop("door_id"),
                         data.pop("location"), data.pop("door_type"), **data)
        await db.doors.insert_one(door)
    job["door_count"] = len(body.doors)
    job.pop("_id", None)
    return job


@api_router.post("/jobs/{job_id}/release")
async def release_job(job_id: str, user=Depends(require_office)):
    res = await db.jobs.update_one({"id": job_id}, {"$set": {"released": True}})
    if not res.matched_count:
        raise HTTPException(404, "Job not found")
    return {"ok": True}


@api_router.get("/doors")
async def list_doors(floor: Optional[str] = None, q: Optional[str] = None,
                     job_id: Optional[str] = None, user=Depends(get_current_user)):
    query = {}
    if floor:
        query["floor"] = floor
    if job_id:
        query["job_id"] = job_id
    if q:
        query["door_id"] = {"$regex": q.strip(), "$options": "i"}
    doors = await db.doors.find(query, {"_id": 0}).sort([("floor", 1), ("door_id", 1)]).to_list(2000)
    counts = await db.files.aggregate([
        {"$match": {"is_deleted": False, "door_id": {"$ne": ""}}},
        {"$group": {"_id": {"$toLower": "$door_id"}, "n": {"$sum": 1}}},
    ]).to_list(1000)
    cmap = {c["_id"]: c["n"] for c in counts}
    for d in doors:
        d["attach_count"] = cmap.get(d["door_id"].lower(), 0)
    return doors


# ---------- Stations ----------

def queue_filter(station: str):
    f = {"stages." + station + ".status": {"$ne": "completed"}}
    if station == "routing":
        f["stages.routing.qc"] = {"$ne": "pass"}
    return f


@api_router.get("/stations/{station}/queue")
async def station_queue(station: str, floor: Optional[str] = None, user=Depends(get_current_user)):
    if station not in STATIONS:
        raise HTTPException(404, "Unknown station")
    if user["role"] == "operator" and user.get("station") != station:
        raise HTTPException(403, "Wrong station")
    jobs = await db.jobs.find({"released": True}, {"_id": 0, "id": 1}).to_list(500)
    job_ids = [j["id"] for j in jobs]
    query = {"job_id": {"$in": job_ids}, **queue_filter(station)}
    if floor:
        query["floor"] = floor
    doors = await db.doors.find(query, {"_id": 0}).sort([("floor", 1), ("door_id", 1)]).to_list(2000)
    out = []
    for d in doors:
        ready = all(d["stages"][p]["status"] == "completed" for p in (PREV[station] or []))
        out.append({**d, "station_ready": ready})
    return out


async def complete_one(door_id: str, station: str, by: str):
    door = await db.doors.find_one({"door_id": door_id})
    if not door:
        raise HTTPException(404, f"Door {door_id} not found")
    job = await db.jobs.find_one({"id": door["job_id"]})
    if not job or not job.get("released"):
        raise HTTPException(400, "Job not released to factory yet")
    if door["stages"][station]["status"] == "completed":
        return door
    for p in PREV[station] or []:
        if door["stages"][p]["status"] != "completed":
            raise HTTPException(400, f"{door_id}: {p.title()} must be completed first")
    if station == "assembly" and not door["stages"]["assembly"].get("photo"):
        raise HTTPException(400, f"{door_id}: photo must be uploaded before completing assembly")
    door["stages"][station].update({"status": "completed", "by": by, "at": now_iso()})
    await db.doors.update_one({"door_id": door_id}, {"$set": {f"stages.{station}": door["stages"][station]}})
    door.pop("_id", None)
    return door


@api_router.post("/doors/{door_id}/stations/{station}/complete")
async def complete_station(door_id: str, station: str, user=Depends(get_current_user)):
    if station not in ["core", "skin", "assembly", "press"]:
        raise HTTPException(400, "Use the QC endpoint for routing")
    if user["role"] == "operator" and user.get("station") != station:
        raise HTTPException(403, "Wrong station")
    return await complete_one(door_id, station, user["name"])


class BatchIn(BaseModel):
    station: str
    door_ids: List[str]


@api_router.post("/doors/batch-complete")
async def batch_complete(body: BatchIn, user=Depends(get_current_user)):
    if body.station not in ["core", "skin", "press"]:
        raise HTTPException(400, "Batch complete not supported for this station")
    if user["role"] == "operator" and user.get("station") != body.station:
        raise HTTPException(403, "Wrong station")
    done, errors = [], []
    for did in body.door_ids:
        try:
            await complete_one(did, body.station, user["name"])
            done.append(did)
        except HTTPException as e:
            errors.append(e.detail)
    return {"completed": done, "errors": errors}


class PhotoIn(BaseModel):
    photo: str


@api_router.post("/doors/{door_id}/photo")
async def upload_photo(door_id: str, body: PhotoIn, user=Depends(get_current_user)):
    if len(body.photo) > 6_000_000:
        raise HTTPException(400, "Photo too large")
    res = await db.doors.update_one({"door_id": door_id},
                                    {"$set": {"stages.assembly.photo": body.photo}})
    if not res.matched_count:
        raise HTTPException(404, "Door not found")
    return {"ok": True}


class QCIn(BaseModel):
    result: str
    notes: str = ""


@api_router.post("/doors/{door_id}/routing/qc")
async def routing_qc(door_id: str, body: QCIn, user=Depends(get_current_user)):
    if user["role"] == "operator" and user.get("station") != "routing":
        raise HTTPException(403, "Wrong station")
    if body.result not in ["pass", "fail"]:
        raise HTTPException(400, "Result must be pass or fail")
    if body.result == "fail" and not body.notes.strip():
        raise HTTPException(400, "Notes are required when a door fails QC")
    door = await db.doors.find_one({"door_id": door_id})
    if not door:
        raise HTTPException(404, "Door not found")
    if door["stages"]["press"]["status"] != "completed":
        raise HTTPException(400, "Press must be completed first")
    st = door["stages"]["routing"]
    st.update({"status": "completed" if body.result == "pass" else "failed",
               "qc": body.result, "notes": body.notes, "by": user["name"], "at": now_iso()})
    await db.doors.update_one({"door_id": door_id}, {"$set": {"stages.routing": st}})
    door["stages"]["routing"] = st
    door.pop("_id", None)
    return door


@api_router.post("/doors/{door_id}/despatch")
async def despatch_door(door_id: str, user=Depends(get_current_user)):
    door = await db.doors.find_one({"door_id": door_id})
    if not door:
        raise HTTPException(404, "Door not found")
    if not (door["stages"]["routing"]["status"] == "completed" and door["stages"]["routing"].get("qc") == "pass"):
        raise HTTPException(400, "Door must pass QC before despatch")
    door["stages"]["despatch"].update({"status": "completed", "by": user["name"], "at": now_iso()})
    await db.doors.update_one({"door_id": door_id}, {"$set": {"stages.despatch": door["stages"]["despatch"]}})
    door.pop("_id", None)
    return door


@api_router.get("/despatch-note")
async def despatch_note(floor: str, user=Depends(get_current_user)):
    doors = await db.doors.find(
        {"floor": floor, "stages.routing.qc": "pass", "stages.despatch.status": {"$ne": "completed"}},
        {"_id": 0}).sort("door_id", 1).to_list(1000)
    return {"floor": floor, "generated_at": now_iso(), "doors": doors}


@api_router.get("/stats")
async def stats(user=Depends(get_current_user)):
    total = await db.doors.count_documents({})
    delivered = await db.doors.count_documents({"stages.despatch.status": "completed"})
    failed = await db.doors.count_documents({"stages.routing.qc": "fail"})
    awaiting_despatch = await db.doors.count_documents(
        {"stages.routing.qc": "pass", "stages.despatch.status": {"$ne": "completed"}})
    return {"total": total, "delivered": delivered, "in_production": total - delivered,
            "qc_failed": failed, "awaiting_despatch": awaiting_despatch}


@api_router.get("/")
async def root():
    return {"message": "MAXX DOORS production API"}


# ---------- Files (drawings & data sheets) ----------

@api_router.post("/files")
async def upload_file(file: UploadFile = File(...), door_id: str = Form(""), job_id: str = Form(""),
                      floor: str = Form(""), user=Depends(get_current_user)):
    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty file")
    if len(data) > 25_000_000:
        raise HTTPException(400, "File too large (25MB max)")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "bin"
    path = f"{APP_NAME}/drawings/{uuid.uuid4()}.{ext}"
    result = await asyncio.to_thread(put_object, path, data, file.content_type or "application/octet-stream")
    doc = {"id": str(uuid.uuid4()), "storage_path": result["path"], "original_filename": file.filename,
           "content_type": file.content_type or "application/octet-stream", "size": result["size"],
           "door_id": door_id.strip(), "job_id": job_id.strip(), "floor": floor.strip(),
           "uploaded_by": user["name"], "created_at": now_iso(), "is_deleted": False}
    await db.files.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.get("/files")
async def list_files(door_id: Optional[str] = None, job_id: Optional[str] = None,
                     floor: Optional[str] = None, user=Depends(get_current_user)):
    q = {"is_deleted": False}
    if door_id:
        q["door_id"] = {"$regex": "^" + door_id.strip() + "$", "$options": "i"}
    if job_id:
        q["job_id"] = job_id
    if floor:
        q["floor"] = {"$regex": floor.strip(), "$options": "i"}
    return await db.files.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@api_router.get("/files/{file_id}/download")
async def download_file(file_id: str, user=Depends(get_current_user)):
    rec = await db.files.find_one({"id": file_id, "is_deleted": False})
    if not rec:
        raise HTTPException(404, "File not found")
    data, ct = await asyncio.to_thread(get_object, rec["storage_path"])
    return Response(content=data, media_type=rec.get("content_type") or ct,
                    headers={"Content-Disposition": f'inline; filename="{rec["original_filename"]}"'})


@api_router.delete("/files/{file_id}")
async def delete_file(file_id: str, user=Depends(get_current_user)):
    res = await db.files.update_one({"id": file_id}, {"$set": {"is_deleted": True}})
    if not res.matched_count:
        raise HTTPException(404, "File not found")
    return {"ok": True}


# ---------- Excel import ----------

HEADER_MAP = {
    "door id": "door_id", "door#": "door_id",
    "location": "location", "building": "floor",
    "leaf height": "leaf_height", "door height": "leaf_height",
    "leaf width 1": "leaf_width_1", "door leaf 1": "leaf_width_1",
    "leaf width 2": "leaf_width_2", "door leaf 2": "leaf_width_2",
    "door thickness": "panel_thickness", "panel thickness": "panel_thickness",
    "door type": "door_type", "fire rating": "fire_rating",
    "core qty leaf 1": "core_qty_1", "core leaf 1": "core_cutting_1",
    "core qty leaf 2": "core_qty_2", "core cutting list leaf 2": "core_cutting_2",
    "stile quantity": "stile_qty", "stiles": "stiles",
    "rail size leaf 1": "rail_1", "rail size leaf 2": "rail_2",
}


def norm_cell(c) -> str:
    return " ".join(str(c).replace("\n", " ").split()).lower() if c is not None else ""


def parse_workbook(data: bytes, filename: str):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    title = str(ws["A1"].value or filename.rsplit(".", 1)[0]).strip()
    rows = list(ws.iter_rows(values_only=True))
    header_idx, headers = None, []
    for i, row in enumerate(rows[:10]):
        norms = [norm_cell(c) for c in row]
        if any(n in ("door id", "door#") for n in norms):
            header_idx, headers = i, norms
            break
    if header_idx is None:
        raise HTTPException(400, "Could not find a header row containing 'DOOR ID' or 'Door#'")
    colmap = {}
    for j, h in enumerate(headers):
        field = HEADER_MAP.get(h)
        if field and field not in colmap.values():
            colmap[j] = field
    doors = []
    for row in rows[header_idx + 1:]:
        rec = {}
        for j, field in colmap.items():
            v = row[j] if j < len(row) else None
            if v is not None and str(v).strip():
                rec[field] = str(v).strip()
        if rec.get("door_id"):
            doors.append(rec)
    if not doors:
        raise HTTPException(400, "No door rows found below the header")
    job_name = title.split(" - ", 1)[1].strip() if " - " in title else title
    return {"job_name": job_name, "source_title": title, "count": len(doors), "doors": doors}


@api_router.post("/import/preview")
async def import_preview(file: UploadFile = File(...), user=Depends(require_office)):
    data = await file.read()
    return await asyncio.to_thread(parse_workbook, data, file.filename or "import.xlsx")


def to_int(v, default=0):
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return default


class ImportConfirm(BaseModel):
    name: str
    client: str = ""
    doors: List[Dict[str, str]]


@api_router.post("/import/confirm")
async def import_confirm(body: ImportConfirm, user=Depends(require_office)):
    if not body.doors:
        raise HTTPException(400, "No doors to import")
    ids = [d["door_id"].strip() for d in body.doors if d.get("door_id")]
    if len(ids) != len(set(ids)):
        raise HTTPException(400, "Duplicate Door IDs in this import")
    existing = await db.doors.find({"door_id": {"$in": ids}}, {"_id": 0, "door_id": 1}).to_list(1000)
    if existing:
        raise HTTPException(400, "Door IDs already exist: " + ", ".join(e["door_id"] for e in existing))
    job = {"id": str(uuid.uuid4()), "name": body.name.strip(), "client": body.client.strip(),
           "released": False, "created_at": now_iso()}
    await db.jobs.insert_one(job)
    for rec in body.doors:
        if not rec.get("door_id"):
            continue
        core_qty = to_int(rec.get("core_qty_1")) + to_int(rec.get("core_qty_2"))
        core_cutting = " / ".join(x for x in [rec.get("core_cutting_1"), rec.get("core_cutting_2")] if x)
        door = make_door(
            job["id"], job["name"], rec.get("floor", "Unassigned"), rec["door_id"].strip(),
            rec.get("location", ""), rec.get("door_type", ""),
            leaf_height=rec.get("leaf_height", ""), leaf_width_1=rec.get("leaf_width_1", ""),
            leaf_width_2=rec.get("leaf_width_2", ""), panel_thickness=rec.get("panel_thickness", ""),
            fire_rating=rec.get("fire_rating", ""), core_qty=core_qty, core_cutting=core_cutting,
            extras={k: v for k, v in rec.items() if k in ("stiles", "rail_1", "rail_2", "stile_qty")},
        )
        await db.doors.insert_one(door)
    job["door_count"] = len(ids)
    job.pop("_id", None)
    return job


# ---------- AI assistant ----------

class ChatIn(BaseModel):
    message: str


async def production_snapshot() -> str:
    doors = await db.doors.find({}, {"_id": 0, "door_id": 1, "floor": 1, "location": 1,
                                     "job_name": 1, "stages": 1}).to_list(1000)
    lines = []
    for d in doors:
        st = d["stages"]
        def m(k):
            return "Y" if st[k]["status"] == "completed" else "-"
        qc = st["routing"].get("qc") or ""
        fail_note = f" QC-FAIL({st['routing'].get('notes','')})" if qc == "fail" else ""
        lines.append(f"{d['door_id']} [{d['floor']}|{d['location']}] core:{m('core')} skin:{m('skin')} "
                     f"asm:{m('assembly')} press:{m('press')} routing:{m('routing')}{qc and '/' + qc}"
                     f" despatch:{m('despatch')}{fail_note}")
    return "\n".join(lines)


@api_router.post("/chat")
async def chat(body: ChatIn, user=Depends(get_current_user)):
    msg = body.message.strip()
    if not msg:
        raise HTTPException(400, "Empty message")
    uid = user["id"]
    await db.chat_messages.insert_one({"user_id": uid, "role": "user", "content": msg, "at": now_iso()})
    history = await db.chat_messages.find({"user_id": uid}, {"_id": 0}).sort("at", -1).to_list(11)
    history.reverse()
    transcript = "\n".join(f"{'User' if h['role'] == 'user' else 'Assistant'}: {h['content']}" for h in history[:-1])
    snapshot = await production_snapshot()
    system = (
        "You are MAXX AI, the production assistant for MAXX DOORS, a fire-rated door factory. "
        "You have LIVE access to the factory's door data below. Stage flags: Y = completed, - = not done. "
        "Workflow order: core and skin (parallel) -> assembly -> press -> routing/QC -> despatch. "
        "Answer concisely, plainly, like a sharp production manager. Use door IDs exactly. "
        "If asked for next actions, respect the workflow order.\n\nLIVE DOOR DATA:\n" + (snapshot or "(no doors yet)")
    )
    prompt = f"Conversation so far:\n{transcript}\n\nUser: {msg}" if transcript else msg

    async def gen():
        full = ""
        try:
            chat_client = LlmChat(api_key=os.environ["EMERGENT_LLM_KEY"],
                                  session_id=f"{uid}-{uuid.uuid4()}", system_message=system
                                  ).with_model("openai", "gpt-5.4")
            async for ev in chat_client.stream_message(UserMessage(text=prompt)):
                if isinstance(ev, TextDelta):
                    full += ev.content
                    yield f"data: {json.dumps({'t': ev.content})}\n\n"
                elif isinstance(ev, StreamDone):
                    break
        except Exception as e:
            logger.error("chat error: %s", e)
            full = "Assistant is unavailable right now — please try again shortly."
            yield f"data: {json.dumps({'t': full})}\n\n"
        yield "data: [DONE]\n\n"
        await db.chat_messages.insert_one({"user_id": uid, "role": "assistant", "content": full, "at": now_iso()})

    return StreamingResponse(gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@api_router.get("/chat/history")
async def chat_history(user=Depends(get_current_user)):
    return await db.chat_messages.find({"user_id": user["id"]},
                                       {"_id": 0, "role": 1, "content": 1, "at": 1}).sort("at", 1).to_list(50)


app.include_router(api_router)
