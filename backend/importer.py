import io
import os
import re
import uuid
import asyncio
import subprocess
import tempfile
import mimetypes

from core import (db, make_door, now_iso, to_int, log_activity, notify,
                  put_object, APP_NAME, get_settings, logger)

SKIP_HEADERS = {"complete", ""}

HEADER_MAP = {
    "door id": "door_id", "door#": "door_id",
    "location": "location", "building": "floor",
    "leaf height": "leaf_height", "door height": "leaf_height",
    "leaf width 1": "leaf_width_1", "door leaf 1": "leaf_width_1",
    "leaf width 2": "leaf_width_2", "door leaf 2": "leaf_width_2",
    "door thickness": "panel_thickness", "panel thickness": "panel_thickness",
    "door type": "door_type", "door schedule type": "door_schedule_type",
    "qty": "qty", "internal door": "internal_door",
    "leaf type single or pair": "leaf_type", "leaf type": "leaf_type_code",
    "panel finish": "panel_finish", "handing": "handing",
    "fire rating": "fire_rating", "frame type": "frame_type",
    "cladding to door": "cladding",
    "vision panel": "vision_panel", "vision panel type": "vision_panel_type",
    "vp size": "vp_size",
    "door grille cutout": "grille_cutout", "grille size": "grille_size",
    "strike prep to frame": "strike_prep",
    "core qty leaf 1": "core_qty_1", "core leaf 1": "core_cutting_1",
    "core qty leaf 2": "core_qty_2", "core cutting list leaf 2": "core_cutting_2",
    "core cutting list leaf 1": "core_cutting_1",
    "skin type": "skin_type", "skin tpye": "skin_type",
    "skin qty leaf 1": "skin_qty_1", "skin cutting list": "skin_cutting_1",
    "skin qty leaf 2": "skin_qty_2", "skin cutting list leaf 2": "skin_cutting_2",
    "skin cutting list leaf 1": "skin_cutting_1",
    "stile quantity": "stile_qty", "stiles": "stiles",
    "rail quantity leaf 1": "rail_qty_1", "rail size leaf 1": "rail_1",
    "rail quantity leaf 2": "rail_qty_2", "rail size leaf 2": "rail_2",
}

PREFIX_MAP = [
    ("hinge qty", "hinge_qty"),
    ("conduit", "conduit"),
    ("door seal prep", "door_seal_prep"),
    ("frame strike height", "frame_strike_height"),
    ("leaf type single", "leaf_type"),
]


def norm_cell(c) -> str:
    return " ".join(str(c).replace("\n", " ").split()).lower() if c is not None else ""


def map_header(h: str):
    if h in SKIP_HEADERS:
        return None
    if h in HEADER_MAP:
        return HEADER_MAP[h]
    for prefix, field in PREFIX_MAP:
        if h.startswith(prefix):
            return field
    return "extra:" + h


def parse_workbook(data: bytes, filename: str):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        header_idx, headers = None, []
        for i, row in enumerate(rows[:10]):
            norms = [norm_cell(c) for c in row]
            if any(n in ("door id", "door#") for n in norms):
                header_idx, headers = i, norms
                break
        if header_idx is None:
            continue
        title = str(ws["A1"].value or filename.rsplit(".", 1)[0]).strip()
        colmap = {}
        for j, h in enumerate(headers):
            field = map_header(h)
            if field and (field.startswith("extra:") or field not in colmap.values()):
                colmap[j] = field
        doors = []
        for row in rows[header_idx + 1:]:
            rec, extras = {}, {}
            for j, field in colmap.items():
                v = row[j] if j < len(row) else None
                if v is None or not str(v).strip():
                    continue
                val = str(v).strip()
                if field.startswith("extra:"):
                    extras[field[6:]] = val
                else:
                    rec[field] = val
            if rec.get("door_id"):
                if extras:
                    rec["_extras"] = extras
                doors.append(rec)
        if doors:
            job_name = title.split(" - ", 1)[1].strip() if " - " in title else title
            return {"job_name": job_name, "source_title": title, "count": len(doors), "doors": doors}
    raise ValueError("Could not find a header row containing 'DOOR ID' or 'Door#'")


def merge_doors(parsed_list):
    merged, order = {}, []
    for parsed in parsed_list:
        for rec in parsed["doors"]:
            did = rec["door_id"]
            if did not in merged:
                merged[did] = {"_job_name": parsed["job_name"], "_extras": {}}
                order.append(did)
            for k, v in rec.items():
                if k == "_extras":
                    merged[did]["_extras"].update(v)
                elif v and not merged[did].get(k):
                    merged[did][k] = v
    return [merged[d] for d in order]


def guess_floor_from_filename(name: str) -> str:
    if re.search(r"L\s?\d+\s*-\s*\d+", name, re.I):
        return ""
    m = re.search(r"\bL(?:EVEL)?[\s_-]*(\d{1,2})\b", name, re.I)
    return f"LEVEL {int(m.group(1))}" if m else ""


def rec_to_door(job_id, job_name, rec):
    data = {k: v for k, v in rec.items() if not k.startswith("_")}
    extras = rec.get("_extras", {})
    return make_door(job_id, job_name, data.pop("floor", "Unassigned"),
                     data.pop("door_id").strip(), data.pop("location", ""),
                     extras=extras, **data)


async def create_jobs_from_records(records, user, released=False):
    levels = {}
    for rec in records:
        levels.setdefault(rec.get("floor", "Unassigned"), []).append(rec)
    created, skipped = [], []
    for floor in sorted(levels):
        recs = levels[floor]
        ids = [r["door_id"].strip() for r in recs]
        existing = await db.doors.find({"door_id": {"$in": ids}}, {"_id": 0, "door_id": 1}).to_list(2000)
        existing_ids = {e["door_id"] for e in existing}
        fresh = [r for r in recs if r["door_id"].strip() not in existing_ids]
        skipped.extend(sorted(existing_ids))
        if not fresh:
            continue
        job_name = fresh[0].get("_job_name") or f"Imported — {floor}"
        job = {"id": str(uuid.uuid4()), "name": job_name, "client": "",
               "released": released, "created_at": now_iso()}
        await db.jobs.insert_one(job)
        seen = set()
        count = 0
        for rec in fresh:
            did = rec["door_id"].strip()
            if did in seen:
                continue
            seen.add(did)
            await db.doors.insert_one(rec_to_door(job["id"], job_name, rec))
            count += 1
        created.append({"id": job["id"], "name": job_name, "floor": floor,
                        "door_count": count, "released": released})
        await log_activity(user["name"], user["role"], "job_imported",
                           f"{job_name} — {count} doors", job_id=job["id"])
    return created, skipped


async def upload_attachment(filename, data, user):
    existing = await db.files.find_one({"original_filename": filename, "is_deleted": False})
    if existing:
        return None
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "bin"
    ct = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    path = f"{APP_NAME}/drawings/{uuid.uuid4()}.{ext}"
    result = await asyncio.to_thread(put_object, path, data, ct)
    doc = {"id": str(uuid.uuid4()), "storage_path": result["path"], "original_filename": filename,
           "content_type": ct, "size": result["size"], "door_id": "", "job_id": "",
           "floor": guess_floor_from_filename(filename), "uploaded_by": user["name"],
           "created_at": now_iso(), "is_deleted": False}
    await db.files.insert_one(doc)
    doc.pop("_id", None)
    return doc


async def run_import(named_files, user, source, run_id=None):
    run_id = run_id or str(uuid.uuid4())
    settings = await get_settings()
    parsed_list, attach_names, errors = [], [], []
    attachments = []
    for name, data in named_files:
        lower = name.lower()
        if lower.endswith((".xlsx", ".xlsm", ".xls")):
            try:
                parsed_list.append(await asyncio.to_thread(parse_workbook, data, name))
            except Exception as e:
                errors.append(f"{name}: {e}")
        else:
            attachments.append((name, data))
    records = merge_doors(parsed_list)
    created, skipped = await create_jobs_from_records(records, user, released=settings.get("auto_release_imports", False))
    for name, data in attachments:
        try:
            doc = await upload_attachment(name, data, user)
            if doc:
                attach_names.append(name)
        except Exception as e:
            errors.append(f"{name}: upload failed ({e})")
    run = {"id": run_id, "source": source, "status": "done", "by": user["name"],
           "file_count": len(named_files), "files": [n for n, _ in named_files],
           "jobs_created": created, "doors_imported": sum(j["door_count"] for j in created),
           "skipped_door_ids": skipped, "attachments": attach_names,
           "errors": errors, "started_at": now_iso(), "finished_at": now_iso()}
    await db.import_runs.update_one({"id": run_id}, {"$set": run}, upsert=True)
    if created:
        await notify("import", f"Import finished — {run['doors_imported']} doors across {len(created)} level(s)",
                     ", ".join(j["name"] for j in created))
    return run


def download_drive_folder(url: str, dest: str):
    result = subprocess.run(["gdown", "--folder", url, "-O", dest],
                            capture_output=True, text=True, timeout=280)
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip().splitlines()[-1] if result.stderr.strip() else "Drive download failed")


async def run_drive_import(run_id: str, url: str, user: dict):
    tmp = tempfile.mkdtemp(prefix="drive_import_")
    try:
        await db.import_runs.update_one({"id": run_id}, {"$set": {"status": "downloading"}})
        await asyncio.to_thread(download_drive_folder, url, tmp)
        named = []
        for root, _, files in os.walk(tmp):
            for f in sorted(files):
                with open(os.path.join(root, f), "rb") as fh:
                    named.append((f, fh.read()))
        if not named:
            raise RuntimeError("No files found in that Drive folder — make sure it is shared as 'Anyone with the link'")
        await db.import_runs.update_one({"id": run_id}, {"$set": {"status": "processing", "file_count": len(named)}})
        await run_import(named, user, "google_drive", run_id=run_id)
    except Exception as e:
        logger.error("drive import failed: %s", e)
        await db.import_runs.update_one(
            {"id": run_id},
            {"$set": {"status": "failed", "errors": [str(e)], "finished_at": now_iso()}})
    finally:
        import shutil
        shutil.rmtree(tmp, ignore_errors=True)
